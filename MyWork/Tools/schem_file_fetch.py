from google.cloud import storage
import os
import time
from google.api_core.exceptions import TooManyRequests
from openpyxl import Workbook
from io import BytesIO

# Initialize the GCS client
client = storage.Client()

def list_blobs_with_prefix(bucket_name, prefix):
    """Lists all the blobs in the bucket that begin with the prefix."""
    bucket = client.bucket(bucket_name)
    blobs = bucket.list_blobs(prefix=prefix)
    return blobs

def upload_excel_to_gcs(bucket_name, file_path, workbook):
    """Uploads an Excel workbook to GCS."""
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(file_path)
    
    # Save the workbook to a bytes buffer
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    blob.upload_from_file(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def process_schema_files(raw_zone_bucket, raw_zone_folder_path, raw_zone_file_path_base):
    blobs = list_blobs_with_prefix(raw_zone_bucket, raw_zone_folder_path)
    file_counter = 0
    schema_file_counter = 0
    
    # Create a new workbook
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove the default sheet created by openpyxl

    current_folder_name = None
    worksheet = None

    for blob in blobs:
        if blob.name.endswith('.schema.csv'):
            # Extract the folder name and prefix
            folder_name = os.path.dirname(blob.name).split('/')[-1].split('-')[1]
            prefix = os.path.basename(blob.name).split('.')[0]
            
            # Create a new sheet for each folder
            if folder_name != current_folder_name:
                current_folder_name = folder_name
                sheet_title = current_folder_name[:31]  # Limit sheet name to 31 characters
                worksheet = workbook.create_sheet(title=sheet_title)
                worksheet.append([f"{folder_name},{prefix}"])
                schema_file_counter = 0

            # Download the content of the schema file
            schema_data = blob.download_as_text()

            # Split the schema data into lines
            lines = schema_data.strip().split('\n')

            # Write data to the current sheet
            for line in lines:
                worksheet.append([line])
            worksheet.append([])  # Adding a new line for separation

            schema_file_counter += 1

            # If we've processed 12 schema files, create a new sheet
            if schema_file_counter == 12:
                file_counter += 1
                raw_zone_file_path = f"{raw_zone_file_path_base}_part{file_counter}.xlsx"
                upload_excel_to_gcs(raw_zone_bucket, raw_zone_file_path, workbook)
                workbook = Workbook()  # Reset workbook
                workbook.remove(workbook.active)  # Remove the default sheet created by openpyxl

    # Upload any remaining data in the workbook
    if schema_file_counter > 0:
        file_counter += 1
        raw_zone_file_path = f"{raw_zone_file_path_base}_part{file_counter}.xlsx"
        upload_excel_to_gcs(raw_zone_bucket, raw_zone_file_path, workbook)

# Configuration
raw_zone_bucket = 'your-raw-zone-bucket'
raw_zone_folder_path = 'your/raw/zone/folder/path'
raw_zone_file_path_base = 'your/raw/zone/file/path/target'

# Execute the processing function
process_schema_files(raw_zone_bucket, raw_zone_folder_path, raw_zone_file_path_base)
